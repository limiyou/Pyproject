"""
需求：
1. 封装一个read_data方法，通过read_data把Excel中的用例读到一个list中。

cases = [
    {
        'id':1
        'title': '注册成功',
        'expected': {"code": 1, "msg": "注册成功"},
        'data': ('python32', '123456', '123456')
    },

    ...
]
2. 封装一个write_data方法，可以把测试结果写入要执行的用例
"""

import openpyxl

class Handle_Excel(object):
    """让读写Excel格式的测试用例更方便
    Attributes:
        filename: Excel文件的路径。
        sheet_name: 要读取的用例在哪个Sheet中，默认是Sheet1
    """
    titles = ('id', 'title', 'data', 'expected')  # 固定的标题

    def __init__(self, filename, sheet_name=None):
        """创建一个HandleExcel对象
        args:
            filename: excel文件的路径
            sheet_name: 要读取的工作表名，如果没有指定默认使用Sheet1当做工作表名
        """
        self.filename = filename
        if sheet_name is None:
            self.sheet_name = 'Sheet1'
        else:
            self.sheet_name = sheet_name

    def read_data(self) -> list:  # 返回值是list
        cases = []  # 用来存储所有用例
        # 一、读数据
        wb = openpyxl.load_workbook(self.filename)  # 获取工作簿对象
        ws = wb[self.sheet_name]
        # ws.rows获取的是所有行的数据
        for i, row in enumerate(ws.rows):  # 遍历所有行   enumerate是Python内置的方法, i是遍历的序号默认从0开始
            # for cell in row:  # 遍历当前行的所有单元格
            # 如果不是第1行，说明是我们要处理的数据
            if i !=0:  # 过滤表头
                # case = {}  # 用来存储一条用例
                # 此时的row是一个元组，存储的是当前行所有cell对象
                # case['id'] = row[0].value  # 把第1个单元格的值赋值给case['id']
                # case['title'] = row[1].value  # 把第2个单元格的值赋值给case['title']
                # case['data'] = row[2].value  # 把第3个单元格的值赋值给case['data']
                # case['expected'] = row[3].value  # 把第4个单元格的值赋值给case['expected']
                # cases.append(case)

                temp = []
                for cell in row:
                    temp.append(cell.value)
                cases.append(dict(zip(self.titles,temp)))
                # cases.append(dict(zip(self.titles, [cell.value for cell in row])))

        return cases

    def write_data(self, row, column, value) -> None:
        """把数据写到指定的单元格

        args:
            row: 行号
            column: 列号
            value:  要写入的值
        returns:
            None
        """
        wb = openpyxl.load_workbook(self.filename)  # 获取一个WorkBook对象
        ws = wb[self.sheet_name]  # 获取一个WorkSheet对象
        ws.cell(row, column, value)
        wb.save(self.filename)

        return None


if __name__ == '__main__':
    excel = Handle_Excel(r'D:\python\pythonProject\autoTest\day3\注册接口.xlsx')
    for case in excel.read_data():
        print(case)

