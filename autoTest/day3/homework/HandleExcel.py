import openpyxl
class Hand_Excel(object):
    titles=('id', 'title', 'data', 'expected')

    def __init__(self,filename,sheet_name=None):
        """
        定义获取的对象
        :param filename: 获取的文件名
        :param sheet_name: 获取的表名，如果没有指定默认使用Sheet1当做工作表名
        """
        self.filename=filename
        if sheet_name is None:
            self.sheet_name='Sheet1'
        else:
            self.sheet_name=sheet_name
    def read_data(self)->list:  #返回值是list
        cases=[]
        wb=openpyxl.load_workbook(self.filename)
        ws=wb[self.sheet_name]
        for i,row in enumerate(ws.rows):
             if i != 0:
                temp = []
                for cell in row:
                    temp.append(cell.value)
                cases.append(dict(zip(self.titles,temp)))
        return  cases

    def write_data(self,row,column,value):
        """
        把测试结果写入到指定单元格内
        :param row: 行数
        :param column: 列数
        :param value: 要写入的值
        :returns: 要返回的值
        """
        wb=openpyxl.load_workbook(self.filename)
        ws=wb[self.sheet_name]
        ws.cell(row,column,value)
        wb.save(self.filename)
        return None

if __name__=="__main__":
    excel=Hand_Excel(r"D:\python\pythonProject\autoTest\day3\homework\登录接口.xlsx")
    for case in excel.read_data():
        print(case)