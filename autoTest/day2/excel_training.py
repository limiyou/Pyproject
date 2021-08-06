from openpyxl import Workbook,load_workbook
from openpyxl.chart import BarChart3D, Reference

#1.创建一个Workbook对象，用来操作Excel
wb=load_workbook(filename=r'D:\python\pythonProject\autoTest\data\成绩单.xlsx')
#2.获取工作表Sheet1对象
ws=wb['Sheet1']
#3.获取指定range的数据
ranges=ws['A1:J10']
#4.输出每个单元格的数据
#4.1遍历所有行
for row in ranges:
    #4.2遍历每行的所有单元格
    for cell in row:
        #输出单元格的值
        print(cell.value,end=',')
    #  每行输出完毕，结尾输出一个；
    print(";")


#二、使用画图功能
# #1.创建一个图表对象
# chart=BarChart3D()
# #2.设置图表格式
# chart.title='成绩统计'
#
# #3.选择图表标题和数据
# titles=Reference(ws,range_string="Sheet1!B2:B10")
# data=Reference(ws,range_string='Sheet1!C2:H10')
# #向图表中添加数据
# chart.add_data(data)
# #设置标题categories
# chart.set_categories(titles)
# #4.把画好的图表添加到指定位置
# ws.add_chart(chart,'M3')
# wb.save('成绩单.xlsx')




chart=BarChart3D()
#设置图表格式
chart.title='成绩统计'
titles=Reference(ws,range_string='Sheet1!B2:B5')
data=Reference(ws,range_string='Sheet1!C2:H5')
#向图表中添加数据
chart.add_data(data)
#设置标题 categories
chart.set_categories(titles)
#把画好的图表添加到指定位置
ws.add_chart(chart,"N10")
wb.save('成绩单.xlsx')
