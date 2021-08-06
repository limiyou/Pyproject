from openpyxl import Workbook,load_workbook
from openpyxl.chart import BarChart3D,Reference
#加载要读取的Excel文件
wb=load_workbook(filename=r"D:\python\pythonProject\autoTest\data\成绩单.xlsx")
#定位要读取的sheet表
ws=wb['Sheet2']
#定位要读取数据的范围
ranges=ws['A1:H10']
#遍历所有行
for row in ranges:
    #遍历每行的单元格
    for cell in row:
        #打印每个单元格的数据，且以,结尾
        print(cell.value, end=',')
    #输出的每行数据，换行展示且以;结尾
    print(";")


#根据获取的表格数据，输出图表


#定义图表对象
chart=BarChart3D()
#定义图表 标题
chart.title='sum of grade'
#获取 横轴要展示的标题
titles=Reference(ws,range_string='Sheet2!B2:B10')
#获取数据
data=Reference(ws,range_string='Sheet2!C2:G10')
#往表格里添加数据
chart.add_data(data)
#设置标题
chart.set_categories(titles)
#设置图表展示的位置
ws.add_chart(chart,'M4')
#保存Excel
wb.save('sum of grade.xlsx')